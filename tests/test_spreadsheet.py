# Writing datasets to CSV, ODS and XLSX.

import zipfile
from pathlib import Path
from xml.dom.minidom import parseString

import pytest

from snapxo.facts.datasets import numbers_dataset, stats_datasets
from snapxo.formats.tables import file_name, safe_name, write_csv, write_ods, write_xlsx


def _dataset() -> dict:
    return {"key": "m", "title": "Messages over time", "columns": ["Month", "Messages"],
            "rows": [["2026-01", 5], ["2026-02", 9]], "chart": "line", "info": "note"}


def test_a_file_name_says_what_it_is():
    assert file_name("Messages over time", "csv", "2026-08-23") == "snapxo-messages-over-time-2026-08-23.csv"
    assert safe_name("Who writes you most?") == "who-writes-you-most-"


def test_csv_starts_with_the_header_row(tmp_path: Path):
    written = write_csv(_dataset(), tmp_path / "a.csv")

    lines = written.read_text(encoding="utf-8-sig").splitlines()
    assert lines[0] == "Month,Messages"
    assert lines[1] == "2026-01,5"


def test_ods_is_a_zip_with_the_mimetype_stored_first(tmp_path: Path):
    written = write_ods([_dataset()], tmp_path / "a.ods")

    with zipfile.ZipFile(written) as archive:
        assert archive.namelist()[0] == "mimetype"
        assert archive.getinfo("mimetype").compress_type == zipfile.ZIP_STORED
        parseString(archive.read("content.xml"))


def test_ods_keeps_numbers_as_numbers(tmp_path: Path):
    written = write_ods([_dataset()], tmp_path / "a.ods")

    content = zipfile.ZipFile(written).read("content.xml").decode("utf-8")
    assert 'office:value-type="float" office:value="5"' in content
    assert 'office:value-type="string"' in content


def test_xlsx_carries_a_real_chart(tmp_path: Path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    written = write_xlsx([_dataset()], tmp_path / "a.xlsx")

    book = load_workbook(written)
    assert book.sheetnames == ["Messages over time"]
    assert len(book["Messages over time"]._charts) == 1


def test_a_sheet_name_excel_would_refuse_is_trimmed(tmp_path: Path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    odd = dict(_dataset(), title="A/B: really quite an extremely long sheet name")
    written = write_xlsx([odd], tmp_path / "a.xlsx")

    name = load_workbook(written).sheetnames[0]
    assert len(name) <= 31
    assert "/" not in name and ":" not in name


def test_every_chart_has_a_dataset_and_an_explanation():
    series = {
        "months": ["2026-01", "2026-02"],
        "year_ticks": [(0, "2026")],
        "messages_per_month": [1, 2], "chat_media_per_month": [0, 1],
        "snaps_sent_per_month": [1, 0], "snaps_received_per_month": [0, 1],
        "friends_per_month": [3, 4], "story_views_per_month": [0, 0],
        "messages_by_hour": [0] * 24, "messages_by_weekday": [0] * 7,
        "type_distribution": [("Text", 3)], "top_senders": [("Alice", "a", 2)],
        "most_interacted": [("Alice", "a", 2)],
    }

    sets = stats_datasets(series)

    assert {s["key"] for s in sets} >= {"messages-over-time", "activity-by-hour", "type-distribution"}
    assert all(s["info"] and isinstance(s["info"], list) for s in sets)
    assert all(len(s["columns"]) == len(s["rows"][0]) for s in sets if s["rows"])


def test_the_message_charts_warn_about_expired_messages():
    series = {"months": ["2026-01", "2026-02"], "year_ticks": [], "messages_per_month": [1, 2],
              "chat_media_per_month": [0, 1], "snaps_sent_per_month": [0, 0],
              "snaps_received_per_month": [0, 0], "friends_per_month": [0, 0],
              "story_views_per_month": [0, 0], "messages_by_hour": [0] * 24,
              "messages_by_weekday": [0] * 7, "type_distribution": [], "top_senders": [],
              "most_interacted": []}

    # info is a list of paragraphs, so the page can break it up
    info = {s["key"]: " ".join(s["info"]) for s in stats_datasets(series)}

    assert "24 hours" in info["messages-over-time"]
    assert "complete" in info["snaps-over-time"]


def test_the_numbers_dataset_mirrors_the_cards():
    dataset = numbers_dataset([("Messages", 5), ("Chats", 2)])

    assert dataset["columns"] == ["Name", "Value"]
    assert dataset["rows"] == [["Messages", 5], ["Chats", 2]]


def test_a_spreadsheet_cell_carries_both_names():
    series = {"months": ["2026-01", "2026-02"], "year_ticks": [], "messages_per_month": [1, 2],
              "chat_media_per_month": [0, 1], "snaps_sent_per_month": [0, 0],
              "snaps_received_per_month": [0, 0], "friends_per_month": [0, 0],
              "story_views_per_month": [0, 0], "messages_by_hour": [0] * 24,
              "messages_by_weekday": [0] * 7, "type_distribution": [],
              "top_senders": [("Alice", "alice_99", 5), ("bob", "", 3)],
              "most_interacted": []}

    rows = {s["key"]: s["rows"] for s in stats_datasets(series)}["who-writes-you-most"]

    assert rows == [["Alice (alice_99)", 5], ["bob", 3]]
